import openpyxl
import unittest
from excel_project import FILE1, FILE2,range_letter


class deg_reg_test(unittest.TestCase):
    file1 = 'Source_file.xlsx'
    file2 = 'date_hour.xlsx'

    def test_date_record_no_cell(self):
        """
        test:write value in no cell
        """
        wb = openpyxl.load_workbook(self.file1)
        sheet = wb.active
        no_value = sheet['A2'].value
        self.assertIsNotNone(no_value)

    def test_date_record_degree_cell(self):
        """
        test:write value in temperature cell
        """
        wb = openpyxl.load_workbook(self.file1)
        sheet = wb.active
        degree_value = sheet['B2'].value
        self.assertIsNotNone(degree_value)

    def test_reset_data_source_file(self):
        """
        test:delete cells source_file
        """
        wb = openpyxl.load_workbook(self.file1)
        sheet = wb.active
        no_value = sheet['A2'].value
        degree_value = sheet['B2'].value
        self.assertIsNone(degree_value, no_value)

    def test_reset_data_date_hour(self):
        """
        test:delete cells date_hour
        """
        wb = openpyxl.load_workbook(self.file2)
        sheet = wb.active
        cells = ['A2', 'B2', 'C2', 'D2']
        for i in cells:
            self.assertIsNone(sheet[i].value)

    def test_Browse_button(self):
        """
        test:browse button(source_file)
        """
        file1_path_split = FILE1.split('/')
        actual_name_file_1 = file1_path_split[len(file1_path_split) - 1]
        expected = 'Source_file.xlsx'
        self.assertEqual(expected, actual_name_file_1)

    def test_Report_file_button(self):
        """
        test:Report file button(final_report)
        """
        file2_path_split = FILE2.split('/')
        actual_name_file_2 = file2_path_split[len(file2_path_split) - 1]
        expected = 'final_report.xlsx'
        self.assertEqual(expected, actual_name_file_2)

    def test_range_letter_expect_pass(self):
        """
        :test: range letter
        """
        start='A'
        stop='B'
        lista = []

        for i in range_letter(start, stop):
            lista.append(i)
        x = 0
        for j in range_letter(start, stop):
            assert j == lista[x]
            x+=1

    def test_range_letter_expect_fail(self):
        """
        :test: range letter
        """
        start='A'
        stop='B'
        lista = []

        for i in range_letter(start, stop):
            lista.append(i)
        x = 1
        for j in range_letter(start, stop):
            assert j == lista[x]
            x+=1

if __name__ == '__main__':
    unittest.main()



