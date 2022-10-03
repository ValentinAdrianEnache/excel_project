"""
Excel automation project
"""
import time
import tkinter as tk
import random
from tkinter.filedialog import askopenfilename
from datetime import datetime
import openpyxl
from tkinter.tix import *

# GUI
WINDOW = Tk()
WINDOW.title('AppX v1.0')
WINDOW.geometry("1000x550")
WINDOW.configure(background="#1D6F42")
background_text = tk.Label(WINDOW,
                           text='KEEP\nCALM\nIT' + "'" + 'S JUST AN\nEXCEL\nFILE',
                           bg='#1D6F42',
                           font=("Arial", 35, "bold"),
                           fg="white")
background_text.place(x=20, y=85)

BUTTON_FILE_1 = tk.Button(
    text="Load",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg='#abf77e',
    fg="black")
BUTTON_FILE_1.place(x=450, y=20)
tip = Balloon()
tip.bind_widget(BUTTON_FILE_1,
                balloonmsg="Load source file")

BUTTON_FILE_2 = tk.Button(
    text="Report",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="#abf77e",
    fg="black")
BUTTON_FILE_2.place(x=550, y=120)
tip.bind_widget(BUTTON_FILE_2,
                balloonmsg="Load report")

BUTTON_DATE_RECORDER = tk.Button(
    text="Rec",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="#abf77e",
    fg="black")
BUTTON_DATE_RECORDER.place(x=450, y=220)
tip.bind_widget(BUTTON_DATE_RECORDER,
                balloonmsg="Register data")

BUTTON_EXECUTION = tk.Button(
    text="Execute",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="#3EB489",
    fg="black")
BUTTON_EXECUTION.place(x=550, y=320)
tip.bind_widget(BUTTON_EXECUTION,
                balloonmsg="Run")


BUTTON_RESET_DATA = tk.Button(
    text="Delete",
    width=15,
    height=3,
    font=("Arial", 15, "bold"),
    bg="#E00201",
    fg="black")
BUTTON_RESET_DATA.place(x=450, y=420)
tip.bind_widget(BUTTON_RESET_DATA,
                balloonmsg="Reset data")

FILE1 = None
FILE2 = None
FILE3 = 'date_hour.xlsx'

WB = openpyxl.load_workbook(FILE3)
SHEET = WB.active

def file_1(event):
    """
    :param event:
    :return:FILE1 name
    """
    global FILE1
    FILE1 = askopenfilename(
        filetypes=[('FILE1', 'source_file.xlsx'),('all files', '*.*')])
    print(FILE1, ' loaded!')


def file_2(event):
    """
        :param event:
        :return:FILE2 name
        """
    global FILE2
    FILE2 = askopenfilename(
        filetypes=[('FILE2', 'report.xlsx'),('all files', '*.*')])
    print(FILE2, ' loaded!')

def date_recorder(event):
    """
    :param event:
    :return: record data
    """

    # write date and hour:
    WB = openpyxl.load_workbook(FILE3)
    SHEET = WB['Sheet']

    date_now = datetime.now()  # current date and hour
    hour_now = date_now.strftime('%H:%M:%S')  # hour format
    day_now = date_now.strftime('%d''-''%m''-''%y')  # day format
    epoch = time.time()  # epoch time in seconds (from 01.01.1970)

    numero_cell = ['A' + str(i) for i in range(1, 22)]
    hour_cell = ['B' + str(i) for i in range(1, 22)]
    date_cell = ['C' + str(i) for i in range(1, 22)]
    epoch_cell = ['D' + str(i) for i in range(1, 22)]

    count = 1
    delta = 6
    wait_sec = 5

    for i in epoch_cell[1:21]:
        if SHEET[i].value is not None:
            time_float = float(SHEET[i].value)
            delta = float(epoch) - time_float
            count += 1
            if count == 21:
                print('Full memory! Press delete for reset data!')

        elif delta > wait_sec:
            SHEET[i].value = epoch
            SHEET[date_cell[count]].value = day_now
            SHEET[hour_cell[count]].value = hour_now
            SHEET[numero_cell[count]].value = str(count) + '.'
            print('Values has been recorded!')
            deg_reg()  # write temperature data
            break
        else:
            print(f'Please wait {wait_sec - int(delta)} seconds until the next record! ')
            break
    WB.save(FILE3)

def deg_reg():
    """
    :return: temperature data
    """
    global FILE1
    WB = openpyxl.load_workbook(FILE1)
    SHEET = WB['Sheet']
    degree = random.randint(18, 22)
    numero_cell = ['A' + str(i) for i in range(1, 22)]
    degree_cell = ['B' + str(i) for i in range(1, 22)]

    count = 0
    for i in numero_cell[1:21]:
        count += 1
        if SHEET[i].value is None:
            SHEET[i].value = str(count) + '.'
            SHEET[degree_cell[count]].value = int(degree)
            break
    WB.save(FILE1)


def range_letter(start, stop):
    """
    :param start: first letter
    :param stop: last letter
    :return: the character that represents the unicode
    """
    return (chr(n) for n in range(ord(start), ord(stop) + 1))

def execute(event):
    """
    :param event:
    :return: create report
    """
    # load and read from file 1
    WB = openpyxl.load_workbook(FILE1)
    SHEET = WB.active
    no_list = []
    t_list = []
    nominal_list = []
    l_tol_list = []
    u_tol_list = []

    for i in range(1, 22):
        no_cell = SHEET['A' + str(i)].value
        no_list.append(no_cell)
        temp_cell = SHEET['B' + str(i)].value
        t_list.append(temp_cell)
        if temp_cell is not None:
            nominal_list.append(20)
            l_tol_list.append(18)
            u_tol_list.append(22)
        else:
            nominal_list.append(None)
            l_tol_list.append(None)
            u_tol_list.append(None)
    # load and read from FILE3
    WB = openpyxl.load_workbook(FILE3)
    SHEET = WB.active
    h_list = []
    d_list = []
    for i in range(1, 22):
        hour_cell = SHEET['B' + str(i)].value
        h_list.append(hour_cell)
        date_cell = SHEET['C' + str(i)].value
        d_list.append(date_cell)

    # write report
    line_no = [str(i) + '2' for i in range_letter("A", "Z")]
    line_temp = [str(i) + '3' for i in range_letter("A", "Z")]
    line_hour = [str(i) + '4' for i in range_letter("A", "Z")]
    line_date = [str(i) + '5' for i in range_letter("A", "Z")]
    nominal_line = [str(i) + '8' for i in range_letter("A", "Z")]
    l_tol_line = [str(i) + '9' for i in range_letter("A", "Z")]
    u_tol_line = [str(i) + '10' for i in range_letter("A", "Z")]

    WB = openpyxl.load_workbook(FILE2)
    SHEET = WB.active


    for i in range(21):
        # start_no_list= start index for each list
        start_no_list = no_list[i]
        start_t_list = t_list[i]
        start_h_list = h_list[i]
        start_d_list = d_list[i]
        start_nom_list = nominal_list[i]
        start_l_tol_list = l_tol_list[i]
        start_u_tol_list = u_tol_list[i]


        SHEET[line_no[i]].value = start_no_list
        SHEET[line_temp[i]].value = start_t_list
        SHEET[line_hour[i]].value = start_h_list
        SHEET[line_date[i]].value = start_d_list
        if i >= 1:
            SHEET[nominal_line[i]].value = start_nom_list
            SHEET[u_tol_line[i]].value = start_l_tol_list
            SHEET[l_tol_line[i]].value = start_u_tol_list
    print('Report has been created! Report path: ', FILE2)
    WB.save(FILE2)


def reset_data(event):
    """
    :param event:
    :return: clear table from all files
    """
    # read and load FILE3
    global FILE3
    WB = openpyxl.load_workbook(FILE3)
    SHEET = WB.active
    # delete columns
    for i in range(4):  # using for loop with range(4) to delete column 1 four times
        SHEET.delete_cols(1)  # when deleted column 1 it's been replaced with column 2 and so on
    # write
    numero = SHEET.cell(row=1, column=1)
    hour = SHEET.cell(row=1, column=2)
    date = SHEET.cell(row=1, column=3)
    time = SHEET.cell(row=1, column=4)

    number_title = 'No.'
    hour_title = 'Hour'
    date_title = 'Date'
    epoch_title = 'Epoch'

    numero.value = number_title
    hour.value = hour_title
    date.value = date_title
    time.value = epoch_title
    WB.save(FILE3)

    global FILE1
    WB = openpyxl.load_workbook(FILE1)
    SHEET = WB.active

    for i in range(2):
        SHEET.delete_cols(1)
        numero = SHEET.cell(row=1, column=1)
        degree = SHEET.cell(row=1, column=2)
        number = 'No.'
        d1 = 'Temperature(°C)'
        numero.value = number
        degree.value = d1
    WB.save(FILE1)
    print('Data has been deleted!')


BUTTON_FILE_1.bind("<Button>", file_1)
BUTTON_FILE_2.bind("<Button>", file_2)
BUTTON_EXECUTION.bind("<Button>", execute)
BUTTON_RESET_DATA.bind("<Button>", reset_data)
BUTTON_DATE_RECORDER.bind("<Button>", date_recorder)
WINDOW.mainloop()
